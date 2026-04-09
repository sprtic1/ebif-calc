"""Tear Sheet Scanner — detects colored highlights in PDFs and extracts text via OCR.

Scans published tear sheet PDFs from the FROM ARCHICAD folder, detects
colored highlight rectangles (yellow, green, blue, pink, purple), OCRs
the text under each highlight, and writes it into the correct manual
column of the project's EBIF SCHEDULES Excel file.

NEVER overwrites cells that already have content — only fills empty cells.

Dependencies: PyMuPDF (fitz), OpenCV, pytesseract, Pillow, numpy
Requires: Tesseract OCR binary installed on the system
"""

import logging
import os
import re
import shutil

import cv2
import fitz  # PyMuPDF
import numpy as np
from PIL import Image

logger = logging.getLogger(__name__)

# Tesseract path (Windows default)
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# PDF render DPI
RENDER_DPI = 300

# Minimum contour area to count as a highlight (in pixels at 300 DPI)
MIN_HIGHLIGHT_AREA = 5000

# Color detection ranges in HSV space
# Each entry: (label, hsv_lower, hsv_upper)
HIGHLIGHT_COLORS = [
    ("yellow",  np.array([20, 40, 220]),  np.array([35, 255, 255])),   # warm yellow, excludes brown product images
    ("green",   np.array([35, 40, 180]),  np.array([60, 255, 255])),   # yellow-green highlights
    ("blue",    np.array([80, 40, 180]),  np.array([130, 255, 255])),   # cyan to blue
    ("pink",    np.array([145, 35, 200]), np.array([165, 255, 255])),   # magenta/pink
    ("purple",  np.array([125, 35, 180]), np.array([145, 255, 255])),   # blue-purple
]

# Color → Excel column mapping (1-indexed)
# Default: Yellow=E(5), Green=F(6), Blue=G(7), Pink=H(8), Purple=I(9)
DEFAULT_COLOR_MAP = {
    "yellow": 5,   # E
    "green":  6,   # F
    "blue":   7,   # G
    "pink":   8,   # H
    "purple": 9,   # I
}

# Tab-specific overrides (same color order, different column meanings)
# The column indices stay the same — Yellow always maps to E, etc.
# The meaning changes per tab but the position doesn't.

# Tear Sheet # column in Excel = C (3), data starts at row 4
TEARSHEET_COL = 3
DATA_START = 4


def check_tesseract():
    """Check if Tesseract OCR is installed and accessible."""
    import pytesseract
    if os.path.exists(TESSERACT_CMD):
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD
    try:
        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False


def scan_tearsheets(project_folder, excel_path, on_progress=None):
    """Scan all tear sheet PDFs and write extracted data to Excel.

    Args:
        project_folder: Full path to the project folder
        excel_path: Full path to the EBIF SCHEDULES Excel file
        on_progress: Optional callback(step, total, pdf_name)

    Returns:
        dict with keys:
            processed: int — total PDFs processed
            updated: int — rows updated in Excel
            skipped: int — PDFs with no highlights or no matching row
            errors: [str] — error messages
            details: [{pdf, tear_sheet, extractions}] — per-PDF results
    """
    import pytesseract
    if os.path.exists(TESSERACT_CMD):
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

    from_archicad = os.path.join(project_folder, 'EBIF', 'EXCEL', 'MASTER', 'FROM ARCHICAD')

    if not os.path.exists(from_archicad):
        return {'processed': 0, 'updated': 0, 'skipped': 0,
                'errors': ['FROM ARCHICAD folder not found'], 'details': []}

    # Find all PDFs
    pdfs = sorted([f for f in os.listdir(from_archicad) if f.lower().endswith('.pdf')])

    if not pdfs:
        return {'processed': 0, 'updated': 0, 'skipped': 0,
                'errors': ['No PDF files found in FROM ARCHICAD folder'], 'details': []}

    # Load Excel to find tear sheet # → row mapping
    from openpyxl import load_workbook
    if not os.path.exists(excel_path):
        return {'processed': 0, 'updated': 0, 'skipped': 0,
                'errors': [f'Excel file not found: {excel_path}'], 'details': []}

    wb = load_workbook(excel_path, keep_vba=True)

    results = {
        'processed': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
        'details': [],
    }

    total = len(pdfs)
    for step, pdf_name in enumerate(pdfs, start=1):
        if on_progress:
            on_progress(step, total, pdf_name)

        pdf_path = os.path.join(from_archicad, pdf_name)
        results['processed'] += 1

        try:
            page_results = process_pdf(pdf_path)

            if not page_results:
                results['skipped'] += 1
                results['details'].append({
                    'pdf': pdf_name, 'tear_sheet': None,
                    'extractions': {}, 'status': 'no_highlights'
                })
                continue

            # Write extractions to Excel
            for pr in page_results:
                ts_num = pr.get('tear_sheet_num', '')
                extractions = pr.get('extractions', {})

                if not ts_num or not extractions:
                    results['skipped'] += 1
                    continue

                # Find matching sheet and row
                written = _write_to_excel(wb, ts_num, extractions)
                if written:
                    results['updated'] += 1
                    results['details'].append({
                        'pdf': pdf_name, 'tear_sheet': ts_num,
                        'extractions': extractions, 'status': 'updated'
                    })
                else:
                    results['skipped'] += 1
                    results['details'].append({
                        'pdf': pdf_name, 'tear_sheet': ts_num,
                        'extractions': extractions, 'status': 'no_match'
                    })

        except Exception as e:
            logger.error("Failed to process %s: %s", pdf_name, e)
            results['errors'].append(f"{pdf_name}: {e}")

    # Save Excel
    try:
        from services.excel_writer import _save_with_retry
        _save_with_retry(wb, excel_path)
    except Exception as e:
        results['errors'].append(f"Excel save failed: {e}")

    return results


def process_pdf(pdf_path):
    """Process a single PDF and extract highlighted text from each page.

    Returns list of {tear_sheet_num, extractions: {color: text}} per page.
    """
    doc = fitz.open(pdf_path)
    page_results = []

    for page_num in range(len(doc)):
        page = doc[page_num]

        # Render page to image at 300 DPI
        mat = fitz.Matrix(RENDER_DPI / 72, RENDER_DPI / 72)
        pix = page.get_pixmap(matrix=mat)
        img_data = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)

        # Convert RGB to BGR for OpenCV
        if pix.n == 4:  # RGBA
            img_bgr = cv2.cvtColor(img_data, cv2.COLOR_RGBA2BGR)
        else:  # RGB
            img_bgr = cv2.cvtColor(img_data, cv2.COLOR_RGB2BGR)

        # Detect colored highlights
        highlights = detect_highlights(img_bgr)

        if not highlights:
            continue

        # Extract tear sheet # from page text
        page_text = page.get_text()
        tear_sheet_num = _extract_tear_sheet_num(page_text, os.path.basename(pdf_path))

        # OCR each highlighted region
        extractions = {}
        for color_label, bbox in highlights:
            text = extract_text_from_region(img_bgr, bbox)
            if text:
                extractions[color_label] = text

        if extractions:
            page_results.append({
                'tear_sheet_num': tear_sheet_num,
                'extractions': extractions,
            })

    doc.close()
    return page_results


def detect_highlights(img_bgr):
    """Detect colored highlight rectangles in an image.

    Returns list of (color_label, (x, y, w, h)) tuples.
    """
    img_hsv = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2HSV)
    results = []

    for color_label, hsv_lower, hsv_upper in HIGHLIGHT_COLORS:
        mask = cv2.inRange(img_hsv, hsv_lower, hsv_upper)

        # Morphological close to fill gaps
        kernel = np.ones((5, 5), np.uint8)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)

        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area >= MIN_HIGHLIGHT_AREA:
                x, y, w, h = cv2.boundingRect(cnt)
                results.append((color_label, (x, y, w, h)))

    return results


def extract_text_from_region(img_bgr, bbox):
    """Crop a region from the image and OCR it.

    Returns cleaned text string, or empty string if OCR fails.
    """
    import pytesseract

    x, y, w, h = bbox
    # Add small padding
    pad = 5
    x = max(0, x - pad)
    y = max(0, y - pad)
    w = min(img_bgr.shape[1] - x, w + 2 * pad)
    h = min(img_bgr.shape[0] - y, h + 2 * pad)

    crop = img_bgr[y:y+h, x:x+w]

    # Convert to grayscale and threshold for better OCR
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # Convert to PIL Image for pytesseract
    pil_img = Image.fromarray(thresh)

    try:
        text = pytesseract.image_to_string(pil_img, config='--psm 6').strip()
        # Clean up: collapse spaces within each line, preserve line breaks
        lines = [re.sub(r'[ \t]+', ' ', line).strip(' |') for line in text.splitlines()]
        lines = [l for l in lines if l]  # remove empty lines
        text = '\n'.join(lines)
        # Strip trailing punctuation artifacts (colons, periods, semicolons, etc.)
        text = re.sub(r'[.:;,!]+$', '', text).strip()
        return text if text and len(text) > 1 else ''
    except Exception as e:
        logger.warning("OCR failed for region: %s", e)
        return ''


def _extract_tear_sheet_num(page_text, filename):
    """Extract the tear sheet number from page text or filename.

    Looks for patterns like "P1.01", "A2.03", "F1.01" etc.
    Falls back to parsing the filename.
    """
    # Try to find tear sheet pattern in page text (e.g., "P1.01", "A2.03")
    match = re.search(r'\b([A-Z]\d+[._]\d+)\b', page_text)
    if match:
        return match.group(1).replace('_', '.')

    # Try filename: e.g., "P1_01_VERELLEN.pdf" → "P1.01"
    name = os.path.splitext(filename)[0]
    match = re.match(r'^([A-Z]\d+)[_.](\d+)', name)
    if match:
        return f"{match.group(1)}.{match.group(2)}"

    # Fallback: just return the filename stem
    return name


def _write_to_excel(wb, tear_sheet_num, extractions):
    """Write extracted data into the correct Excel cell.

    Scans all schedule sheets for a matching Tear Sheet # in column C.
    Only writes to empty cells — NEVER overwrites existing data.

    Returns True if at least one cell was written.
    """
    written = False

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Scan column C (Tear Sheet #) for matching value
        for row in range(DATA_START, (ws.max_row or DATA_START) + 1):
            cell_val = ws.cell(row=row, column=TEARSHEET_COL).value
            if cell_val is None:
                continue

            cell_str = str(cell_val).strip()
            if cell_str == tear_sheet_num or cell_str.replace('.', '_') == tear_sheet_num.replace('.', '_'):
                # Found matching row — write extractions to manual columns
                for color_label, text in extractions.items():
                    col_idx = DEFAULT_COLOR_MAP.get(color_label)
                    if col_idx is None:
                        continue

                    existing = ws.cell(row=row, column=col_idx).value
                    if existing is not None and str(existing).strip():
                        logger.info("Skipping %s row %d col %d — already has data: '%s'",
                                    sheet_name, row, col_idx, str(existing)[:30])
                        continue

                    cell = ws.cell(row=row, column=col_idx, value=text)
                    from openpyxl.styles import Alignment
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    written = True
                    logger.info("Wrote '%s' to %s row %d col %d (color=%s)",
                                text[:30], sheet_name, row, col_idx, color_label)

                return written  # Found the row, done with this tear sheet

    logger.warning("No matching row found for tear sheet '%s'", tear_sheet_num)
    return False


def publish_tearsheets(port, project_folder):
    """Trigger the Archicad publisher set to export tear sheet PDFs.

    Uses Tapir's PublishPublisherSet command.
    """
    import requests

    output_path = os.path.join(project_folder, 'EBIF', 'EXCEL', 'MASTER', 'FROM ARCHICAD')
    os.makedirs(output_path, exist_ok=True)

    url = f"http://localhost:{port}"
    payload = {
        "command": "API.ExecuteAddOnCommand",
        "parameters": {
            "addOnCommandId": {
                "commandNamespace": "TapirCommand",
                "commandName": "PublishPublisherSet"
            },
            "addOnCommandParameters": {
                "publisherSetName": "PDF - TEAR SHEETS",
                "outputPath": output_path.replace("\\", "/")
            }
        }
    }

    resp = requests.post(url, json=payload, timeout=120)
    data = resp.json()

    if not data.get("succeeded", True):
        err = data.get("error", {})
        raise RuntimeError(f"Publish failed: {err.get('message', err)}")

    return output_path
