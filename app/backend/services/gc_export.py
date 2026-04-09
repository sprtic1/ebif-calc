"""GC Export — generates a General Contractor Excel package.

Reads the project's EBIF SCHEDULES file and creates a clean,
formatted Excel file with only the columns a GC needs:
Item, Location, Manufacturer, Model #, Size, Finish, Notes, Qty.

No cost, vendor, or sourcing columns. EID branded.
"""

import logging
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook as load_wb

logger = logging.getLogger(__name__)

# EID Brand
OLIVE = "868C54"
SAGE = "F0F2E8"
WHITE = "FFFFFF"
WARM_GRAY = "737569"

HEADER_FONT = Font(name="Lato", bold=True, color=WHITE, size=10)
HEADER_FILL = PatternFill(start_color=OLIVE, end_color=OLIVE, fill_type="solid")
BODY_FONT = Font(name="Arial Narrow", color="2C2C2C", size=10)
ALT_FILL = PatternFill(start_color=SAGE, end_color=SAGE, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
TITLE_FONT = Font(name="Lato", bold=True, color=OLIVE, size=14)
SUBTITLE_FONT = Font(name="Arial Narrow", color=WARM_GRAY, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Source columns in the Master Excel (1-indexed)
# GC output columns and where to read them from
GC_COLUMNS = [
    {"label": "Item", "source_col": 3},          # Tear Sheet # as item identifier
    {"label": "Location", "source_col": 4},
    {"label": "Manufacturer", "source_col": 5},
    {"label": "Model #", "source_col": 6},
    {"label": "Size", "source_col": 7},
    {"label": "Finish", "source_col": 8},
    {"label": "Notes", "source_col": 9},
    {"label": "Qty", "source_col": 2},
]

DATA_START = 4  # First data row in source Excel

# Schedule tabs
SCHEDULE_TABS = {
    'appliances': 'Appliances',
    'bath_accessories': 'Bath Accessories',
    'cabinetry_hardware': 'Cabinetry Hardware',
    'cabinetry_inserts': 'Cabinetry Inserts',
    'cabinetry_style': 'Cabinetry Style & Species',
    'countertops': 'Countertops',
    'decorative_lighting': 'Decorative Lighting',
    'door_hardware': 'Door Hardware',
    'flooring': 'Flooring',
    'furniture': 'Furniture',
    'lighting_electrical': 'Lighting & Electrical',
    'plumbing': 'Plumbing',
    'shower_glass_mirrors': 'Shower Glass & Mirrors',
    'specialty_equipment': 'Specialty Equipment',
    'surface_finishes': 'Surface Finishes',
    'tile': 'Tile',
}


def generate_gc_package(project):
    """Generate the GC Excel package from the project's Master Excel.

    Args:
        project: Project dict with folder_location, project_name, client_name

    Returns:
        dict with:
            path: str — full path to the generated file
            filename: str — just the filename
            tabs: int — number of schedule tabs included
            rows: int — total data rows
    """
    from services.template import get_excel_path

    source_path = get_excel_path(project)
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"Source Excel not found: {source_path}")

    # Open source workbook
    src_wb = load_wb(source_path, read_only=True, data_only=True)

    # Create GC workbook
    gc_wb = Workbook()
    gc_wb.remove(gc_wb.active)  # Remove default sheet

    project_name = project.get('project_name', 'Project')
    client_name = project.get('client_name', '')
    date_str = datetime.now().strftime('%Y-%m-%d')

    total_rows = 0
    tabs_included = 0

    for sid, tab_name in SCHEDULE_TABS.items():
        src_ws = _find_sheet(src_wb, tab_name)
        if src_ws is None:
            continue

        # Read data rows — check if tab has any data
        rows = []
        for row_cells in src_ws.iter_rows(min_row=DATA_START, max_col=11):
            # Check if row has data (col 2=QTY or col 5=manufacturer)
            has_data = False
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip() not in ('', '#REF!', '0'):
                    has_data = True
                    break
            if not has_data:
                break

            row_data = {}
            for gc_col in GC_COLUMNS:
                src_idx = gc_col["source_col"] - 1  # 0-indexed in tuple
                val = row_cells[src_idx].value if src_idx < len(row_cells) else None
                if val is None or str(val).strip() in ('#REF!', ''):
                    val = ''
                row_data[gc_col["label"]] = val
            rows.append(row_data)

        if not rows:
            continue

        # Create tab in GC workbook
        gc_ws = gc_wb.create_sheet(title=tab_name[:31])
        tabs_included += 1

        # Title rows
        gc_ws.cell(row=1, column=1, value=tab_name).font = TITLE_FONT
        gc_ws.cell(row=2, column=1, value=f"{project_name} — {date_str}").font = SUBTITLE_FONT

        # Header row at row 4
        header_labels = [c["label"] for c in GC_COLUMNS]
        for col_idx, label in enumerate(header_labels, start=1):
            cell = gc_ws.cell(row=4, column=col_idx, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

        # Data rows from row 5
        for i, row_data in enumerate(rows):
            row_num = 5 + i
            fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
            for col_idx, gc_col in enumerate(GC_COLUMNS, start=1):
                val = row_data.get(gc_col["label"], '')
                cell = gc_ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        total_rows += len(rows)

        # Column widths
        widths = [15, 20, 20, 20, 15, 15, 30, 8]
        for col_idx, w in enumerate(widths, start=1):
            gc_ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Freeze header
        gc_ws.freeze_panes = "A5"

    src_wb.close()

    if tabs_included == 0:
        raise ValueError("No schedule data found — nothing to export")

    # Save GC file
    folder = project.get('folder_location', '')
    output_dir = os.path.join(folder, 'EBIF', 'EXCEL')
    os.makedirs(output_dir, exist_ok=True)

    filename = f"{client_name} - GC PACKAGE.xlsx" if client_name else "GC PACKAGE.xlsx"
    output_path = os.path.join(output_dir, filename)

    gc_wb.save(output_path)
    logger.info("GC package: %s (%d tabs, %d rows)", filename, tabs_included, total_rows)

    return {
        'path': output_path,
        'filename': filename,
        'tabs': tabs_included,
        'rows': total_rows,
    }


def _find_sheet(wb, schedule_name):
    """Find sheet by name, handling emoji prefixes."""
    for sheet_name in wb.sheetnames:
        if sheet_name == schedule_name:
            return wb[sheet_name]
        stripped = sheet_name
        while stripped and (not stripped[0].isascii() or stripped[0] == ' '):
            stripped = stripped[1:]
        if stripped == schedule_name:
            return wb[sheet_name]
        if sheet_name.endswith(schedule_name):
            return wb[sheet_name]
    return None
