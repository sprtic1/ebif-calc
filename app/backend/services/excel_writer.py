"""EBIF Master Template Excel Writer — writes Archicad data into the template.

All Archicad data is written starting at column N (14) onward:
  N = EBIF UID, O = QTY, P = Tear Sheet #, Q = Location,
  R+ = Element Type, Library Part Name, Layer, Classifications, etc.

Columns A–D contain formulas that reference the Archicad data:
  A =N{row}, B =O{row}, C =P{row}, D =Q{row}

Columns E–K are MANUAL (never touched). L–M are empty (never touched).
Table of Contents tab is never modified.

Row 4 = header row. Data starts at row 5.
On refresh: backup file, check lock, clear A–D + N onward, rewrite, validate.

Robustness:
- Creates timestamped backup before every write (keeps last 5)
- Checks for file lock before writing
- Removes Excel Table objects (handles both str and Table types)
- Retries save up to 3 times with 2s delay
- Per-category error isolation — failures don't abort remaining categories
- Validates first row after write

Uses openpyxl with keep_vba=True to preserve macros.
"""

import glob
import logging
import os
import shutil
import time
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# EID Brand Colors
OLIVE = "868C54"
WHITE = "FFFFFF"

# Styling
BODY_FONT = Font(name="Arial Narrow", color="2C2C2C", size=10)
HEADER_FONT = Font(name="Lato", bold=True, color=WHITE, size=10)
HEADER_FILL = PatternFill(start_color=OLIVE, end_color=OLIVE, fill_type="solid")
LOCKED_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
LOCKED_ALT_FILL = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

# Column width: minimum 50 units for EBIF UID (col N), 15 units for others
MIN_WIDTH_UID = 50
MIN_WIDTH_DEFAULT = 15
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

HEADER_ROW = 3
DATA_START = 4

# Archicad data start column = 2 columns after last manual column.
# Most tabs: last manual = K(11), so AC starts at N(14).
# Furniture: last manual = L(12), so AC starts at O(15).
# Decorative Lighting: last manual = M(13), so AC starts at P(16).
AC_START_DEFAULT = 14  # column N — for 13 standard tabs
AC_START_OVERRIDES = {
    'furniture': 15,            # column O
    'decorative_lighting': 16,  # column P
}

# Gap column width: 250px ≈ 36 Excel units
GAP_COL_WIDTH = 36

# Core 4 Archicad fields (header label, data key)
CORE_FIELDS = [
    ("EBIF UID", "EBIF UID"),
    ("QTY", "Qty"),
    ("Tear Sheet #", "TEAR SHEET #"),
    ("Location", "Location"),
]

# Columns A–D get formulas referencing the core 4 at AC_START+0..+3
FORMULA_COLS = [1, 2, 3, 4]  # A, B, C, D

# Tabs to skip
SKIP_TABS = {"Table of Contents"}

MAX_RETRIES = 3
RETRY_DELAY = 2  # seconds
MAX_BACKUPS = 5


class FileLockError(Exception):
    """Raised when the Excel file is open/locked by another process."""
    pass


def write_to_master(
    xlsm_path,
    schedules,
    schedule_defs,
    on_progress=None,
):
    """Write Archicad data into the project's EBIF schedule file.

    Args:
        xlsm_path: Full path to the project's .xlsm file
        schedules: dict mapping schedule_id -> list of row dicts
        schedule_defs: list of schedule definition dicts (with resolved column info)
        on_progress: Optional callback(step, total, schedule_name, items_so_far, items_total)

    Returns:
        dict with keys:
            counts: {schedule_id: int}
            failed: [schedule_id, ...] — categories that failed (if any)
            warnings: [str, ...] — validation warnings
    """
    if not os.path.exists(xlsm_path):
        raise FileNotFoundError(f"Excel file not found: {xlsm_path}")

    # Check if file is locked
    _check_file_lock(xlsm_path)

    # Create timestamped backup
    _create_backup(xlsm_path)

    wb = load_workbook(xlsm_path, keep_vba=True)
    result = {}
    failed = []
    warnings = []
    total_steps = len(schedule_defs)
    total_items = sum(len(schedules.get(s['id'], [])) for s in schedule_defs)
    items_so_far = 0

    for step, sdef in enumerate(schedule_defs, start=1):
        sid = sdef['id']
        sname = sdef['name']
        rows = schedules.get(sid, [])

        if on_progress:
            on_progress(step, total_steps, sname, items_so_far, total_items)

        items_so_far += len(rows)

        if not rows:
            result[sid] = 0
            continue

        try:
            _write_schedule(wb, sdef, rows)
            result[sid] = len(rows)
            logger.info("Wrote %d rows for '%s'", len(rows), sname)
        except Exception as e:
            logger.error("Failed to write '%s': %s", sname, e)
            failed.append(sid)
            result[sid] = 0

    # Save with retry logic (handles file-in-use by Excel)
    _save_with_retry(wb, xlsm_path)

    # Validate first row of each written schedule
    try:
        wb_check = load_workbook(xlsm_path, keep_vba=True, data_only=True)
        for sdef in schedule_defs:
            sid = sdef['id']
            if result.get(sid, 0) > 0:
                rows = schedules.get(sid, [])
                w = _validate_first_row(wb_check, sdef, rows)
                if w:
                    warnings.append(w)
    except Exception as e:
        logger.warning("Validation read-back failed: %s", e)

    return {'counts': result, 'failed': failed, 'warnings': warnings}


def _get_ac_start(sid):
    """Return the Archicad data start column for a given schedule ID."""
    return AC_START_OVERRIDES.get(sid, AC_START_DEFAULT)


def _write_schedule(wb, sdef, rows):
    """Write a single schedule's data into its sheet."""
    sname = sdef['name']
    sid = sdef['id']

    ws = _find_sheet(wb, sname)
    if ws is None:
        raise ValueError(f"Sheet not found for '{sname}'")

    if any(skip in ws.title for skip in SKIP_TABS):
        return

    ac_start = _get_ac_start(sid)

    # Remove Excel Table objects to prevent repair errors
    _remove_tables(ws)

    # Remove legacy button objects (REFRESH APPLIANCES, etc.) from rows 1-3
    _remove_legacy_buttons(ws)

    # Build reference labels
    archicad_labels = sdef.get('_archicad_col_labels', [])
    core_keys = {f[1] for f in CORE_FIELDS}
    ref_labels = [lbl for lbl in archicad_labels if lbl not in core_keys]
    total_ac_cols = 4 + len(ref_labels)

    # Clear old data
    _clear_data(ws, total_ac_cols, ac_start)

    # Write headers at row 4 starting at ac_start
    all_headers = [f[0] for f in CORE_FIELDS] + ref_labels
    for j, header in enumerate(all_headers):
        col = ac_start + j
        cell = ws.cell(row=HEADER_ROW, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Formula source columns: ac_start+0, +1, +2, +3
    formula_src = [ac_start + i for i in range(4)]

    # Write data rows
    for i, row_data in enumerate(rows):
        row_num = DATA_START + i

        # Core 4 at ac_start, +1, +2, +3
        _write_cell(ws, row_num, ac_start + 0, row_data.get('EBIF UID', ''), i)
        _write_cell(ws, row_num, ac_start + 1, row_data.get('Qty', 1), i)
        _write_cell(ws, row_num, ac_start + 2, row_data.get('TEAR SHEET #', ''), i)
        _write_cell(ws, row_num, ac_start + 3, row_data.get('Location', ''), i)

        # Reference columns after core 4
        for j, lbl in enumerate(ref_labels):
            col = ac_start + 4 + j
            _write_cell(ws, row_num, col, row_data.get(lbl, ''), i)

        # Formulas in A–D referencing the core 4 columns
        for dest_col, src_col in zip(FORMULA_COLS, formula_src):
            src_letter = get_column_letter(src_col)
            cell = ws.cell(row=row_num, column=dest_col, value=f"={src_letter}{row_num}")
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = LIGHT_GRAY_FILL

    # Set gap column widths (2 empty cols between manual and Archicad data)
    _set_gap_column_widths(ws, ac_start)

    # Auto-fit column widths for Archicad columns — don't touch manual columns
    _autofit_archicad_columns(ws, all_headers, rows, ref_labels, ac_start)


def _check_file_lock(xlsm_path):
    """Check if the file is locked by another process (e.g. Excel)."""
    lock_file = os.path.join(os.path.dirname(xlsm_path), '~$' + os.path.basename(xlsm_path))
    if os.path.exists(lock_file):
        raise FileLockError(
            "Please close the Excel file and try again."
        )
    # Also try opening for write to confirm
    try:
        with open(xlsm_path, 'r+b'):
            pass
    except PermissionError:
        raise FileLockError(
            "Please close the Excel file and try again."
        )


def _create_backup(xlsm_path):
    """Create a timestamped backup, keeping only the last MAX_BACKUPS."""
    backup_dir = os.path.join(os.path.dirname(xlsm_path), '_backups')
    os.makedirs(backup_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M')
    base = os.path.splitext(os.path.basename(xlsm_path))[0]
    backup_name = f"{base}_{timestamp}.xlsm"
    backup_path = os.path.join(backup_dir, backup_name)

    shutil.copy2(xlsm_path, backup_path)
    logger.info("Backup created: %s", backup_path)

    # Prune old backups — keep only the last MAX_BACKUPS
    pattern = os.path.join(backup_dir, f"{base}_*.xlsm")
    backups = sorted(glob.glob(pattern), key=os.path.getmtime)
    while len(backups) > MAX_BACKUPS:
        old = backups.pop(0)
        os.remove(old)
        logger.info("Removed old backup: %s", old)


def _save_with_retry(wb, xlsm_path):
    """Save the workbook with retry logic for file-in-use scenarios."""
    last_error = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            wb.save(xlsm_path)
            logger.info("Saved EBIF Master Template: %s", xlsm_path)
            return
        except PermissionError as e:
            last_error = e
            if attempt < MAX_RETRIES:
                logger.warning("Save attempt %d failed (file locked), retrying in %ds...",
                               attempt, RETRY_DELAY)
                time.sleep(RETRY_DELAY)
            else:
                raise FileLockError(
                    "Please close the Excel file and try again."
                ) from e
        except Exception as e:
            last_error = e
            if attempt < MAX_RETRIES:
                logger.warning("Save attempt %d failed: %s, retrying in %ds...",
                               attempt, e, RETRY_DELAY)
                time.sleep(RETRY_DELAY)
            else:
                raise RuntimeError(f"Excel save failed after {MAX_RETRIES} attempts: {e}") from e


def _validate_first_row(wb, sdef, rows):
    """Read back the first data row and verify it matches what was written.

    Returns a warning string if validation fails, or None if OK.
    """
    if not rows:
        return None

    sname = sdef['name']
    ws = _find_sheet(wb, sname)
    if ws is None:
        return f"Validation: sheet not found for '{sname}'"

    ac_start = _get_ac_start(sdef.get('id', ''))
    expected_uid = rows[0].get('EBIF UID', '')
    actual_uid = ws.cell(row=DATA_START, column=ac_start).value

    if str(actual_uid or '').strip() != str(expected_uid or '').strip():
        msg = f"Validation mismatch in '{sname}': expected EBIF UID '{expected_uid}', got '{actual_uid}'"
        logger.warning(msg)
        return msg

    return None


def _remove_tables(ws):
    """Remove all Excel Table objects from a worksheet.

    Uses the TableList's own .clear() method to preserve the correct
    container type. Setting ws._tables = [] (plain list) breaks
    openpyxl's save which calls ws.tables.values().
    """
    tables = getattr(ws, '_tables', None)
    if not tables:
        return

    names = []
    for t in tables.values() if hasattr(tables, 'values') else tables:
        if isinstance(t, str):
            names.append(t)
        elif hasattr(t, 'name'):
            names.append(t.name)
        else:
            names.append(repr(t))

    if hasattr(tables, 'clear'):
        tables.clear()
    else:
        ws._tables = type(tables)() if callable(type(tables)) else tables

    if names:
        logger.info("Removed %d Excel Table(s) from '%s': %s",
                     len(names), ws.title, ', '.join(names))


def _remove_legacy_buttons(ws):
    """Remove legacy VBA button objects from rows 1-3.

    These are REFRESH APPLIANCES / REFRESH ALL buttons from the old
    Power Query system. They live in VML legacy drawings. Clearing
    the legacy_drawing reference removes them from the sheet.
    Also clears any cell content in rows 1-3 from column L onward
    (old PQ-related formulas like #REF!).
    """
    if ws.legacy_drawing:
        logger.info("Removed legacy drawing (buttons) from '%s'", ws.title)
        ws.legacy_drawing = None

    # Clear stray content in rows 1-3 from column L onward
    max_col = ws.max_column or 1
    for row in range(1, 4):
        for col in range(12, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.value = None


def _set_gap_column_widths(ws, ac_start):
    """Set the 2 empty gap columns before Archicad data to 250px (~36 units).

    Gap columns are ac_start-2 and ac_start-1:
      Most tabs: L(12) and M(13)
      Furniture: M(13) and N(14)
      Decorative Lighting: N(14) and O(15)
    """
    for col_idx in (ac_start - 2, ac_start - 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = GAP_COL_WIDTH


def _autofit_archicad_columns(ws, all_headers, rows, ref_labels, ac_start):
    """Auto-fit column widths for Archicad data columns (ac_start onward).

    Does NOT change any column widths for columns before ac_start.
    First column (EBIF UID) gets a minimum width of 50 units (~350px).
    Other columns auto-fit based on content with a minimum of 15 units.
    """
    for j, header in enumerate(all_headers):
        col_idx = ac_start + j
        col_letter = get_column_letter(col_idx)

        # Calculate max content width: header + data
        max_len = len(str(header))
        for row_data in rows[:100]:  # sample first 100 rows for speed
            key = header
            # Map header display name back to data key
            if j < 4:
                key = CORE_FIELDS[j][1]
            val = row_data.get(key, '')
            if val is None:
                val = ''
            max_len = max(max_len, len(str(val)))

        # Convert to width units (roughly 1 char = 1.1 units) + padding
        width = max(max_len * 1.1 + 4, MIN_WIDTH_DEFAULT)

        # Column N (EBIF UID) gets a minimum of 50 units
        if j == 0:
            width = max(width, MIN_WIDTH_UID)

        ws.column_dimensions[col_letter].width = width


def _find_sheet(wb, schedule_name):
    """Find a sheet by schedule name, handling emoji prefixes."""
    if schedule_name in wb.sheetnames:
        return wb[schedule_name]

    for sheet_name in wb.sheetnames:
        stripped = sheet_name
        while stripped and (not stripped[0].isascii() or stripped[0] == ' '):
            stripped = stripped[1:]
        after_emoji = sheet_name[2:] if len(sheet_name) > 2 else sheet_name

        if stripped == schedule_name or after_emoji == schedule_name:
            return wb[sheet_name]
        if sheet_name.endswith(schedule_name):
            return wb[sheet_name]

    truncated = schedule_name[:31]
    if truncated in wb.sheetnames:
        return wb[truncated]

    return None


def _write_cell(ws, row, col, value, row_index):
    """Write a single cell with Archicad styling."""
    if value is None:
        value = ''
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.fill = LOCKED_ALT_FILL if row_index % 2 == 0 else LOCKED_FILL


def _clear_data(ws, total_ac_cols, ac_start):
    """Clear old Archicad data from row 5 downward.

    Clears:
      - A–D (formulas)
      - Gap columns (from col 12 up to ac_start-1) — clears stray data
      - ac_start onward (Archicad data)
    NEVER touches manual columns (E–K, or E–L for Furniture, E–M for Lighting).
    """
    max_row = ws.max_row or DATA_START
    if max_row < DATA_START:
        return

    max_ac_col = ac_start + total_ac_cols
    furthest = max(max_ac_col, ws.max_column + 1) if ws.max_column else max_ac_col

    for row_num in range(DATA_START, max_row + 1):
        # Clear A–D (formula columns)
        for col in range(1, 5):
            ws.cell(row=row_num, column=col).value = None
        # Clear gap columns (between last manual col and ac_start)
        for col in range(12, ac_start):
            ws.cell(row=row_num, column=col).value = None
        # Clear Archicad data columns
        for col in range(ac_start, furthest):
            ws.cell(row=row_num, column=col).value = None
