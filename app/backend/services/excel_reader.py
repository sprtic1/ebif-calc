"""Excel Reader — reads schedule data from the project's EBIF SCHEDULES file.

Provides two levels of detail:
- read_excel_counts(project): simple {schedule_id: count} for tile numbers
- read_excel_details(project): enriched per-schedule data with row detail,
  completeness metrics, and manufacturer fill status for color-coded tiles
"""

import logging
import os

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# EBIF UID column varies per tab (2 cols after last manual column)
AC_START_DEFAULT = 14  # column N — most tabs
AC_START_OVERRIDES = {
    'furniture': 15,            # column O
    'decorative_lighting': 16,  # column P
}
DATA_START = 4
MANUFACTURER_COL = 5  # column E — first manual col (Manufacturer for most tabs)

# Schedule tab name mapping (matches schedules.json IDs)
SCHEDULE_TABS = {
    'appliances': 'Appliances',
    'bath_accessories': 'Bath Accessories',
    'cabinetry_hardware': 'Cabinetry Hardware',
    'cabinetry_inserts': 'Cabinetry Inserts',
    'cabinetry_style': 'Cabinetry Style & Species',
    'countertops': 'Countertops',
    'decorative_lighting': 'Decorative Lighting',
    'door_hardware': 'Door Hardware',
    'flooring': 'Flooring',
    'furniture': 'Furniture',
    'lighting_electrical': 'Lighting & Electrical',
    'plumbing': 'Plumbing',
    'shower_glass_mirrors': 'Shower Glass & Mirrors',
    'specialty_equipment': 'Specialty Equipment',
    'surface_finishes': 'Surface Finishes',
    'tile': 'Tile',
}


def read_excel_counts(project):
    """Read row counts per schedule tab from the project's Excel file.

    Returns dict mapping schedule_id -> count, or None if file not found.
    """
    from services.template import get_excel_path
    xlsm_path = get_excel_path(project)

    if not os.path.exists(xlsm_path):
        return None

    try:
        wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Could not read Excel file: %s", e)
        return None

    counts = {}
    for sid, tab_name in SCHEDULE_TABS.items():
        ws = _find_sheet(wb, tab_name)
        if ws is None:
            counts[sid] = 0
            continue

        ac_col = AC_START_OVERRIDES.get(sid, AC_START_DEFAULT)
        count = 0
        for row in ws.iter_rows(min_row=DATA_START, min_col=ac_col, max_col=ac_col, values_only=True):
            if row[0] is not None and str(row[0]).strip():
                count += 1
            else:
                break
        counts[sid] = count

    wb.close()
    return counts


def read_excel_details(project):
    """Read enriched schedule data from the project's Excel file.

    Returns dict mapping schedule_id -> {
        count: int,
        complete: int,       # rows with manufacturer (col E) filled
        incomplete: int,     # rows with items but no manufacturer
        rows: [{ebif_uid, tear_sheet, location, manufacturer}, ...]
    }
    Or None if file not found.
    """
    from services.template import get_excel_path
    xlsm_path = get_excel_path(project)

    if not os.path.exists(xlsm_path):
        return None

    try:
        wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    except Exception as e:
        logger.warning("Could not read Excel file: %s", e)
        return None

    details = {}
    for sid, tab_name in SCHEDULE_TABS.items():
        ws = _find_sheet(wb, tab_name)
        if ws is None:
            details[sid] = {'count': 0, 'complete': 0, 'incomplete': 0, 'rows': []}
            continue

        ac_col = AC_START_OVERRIDES.get(sid, AC_START_DEFAULT)
        rows_data = []
        complete = 0
        incomplete = 0

        for row_cells in ws.iter_rows(min_row=DATA_START, max_col=max(ac_col + 4, MANUFACTURER_COL + 1)):
            # Check EBIF UID (ac_col, 0-indexed in the row tuple is ac_col-1)
            uid_val = row_cells[ac_col - 1].value if len(row_cells) >= ac_col else None
            if uid_val is None or not str(uid_val).strip():
                break  # end of data

            # Read key columns
            # Col C (3) = Tear Sheet #, Col D (4) = Location, Col E (5) = Manufacturer
            tear_sheet = _cell_val(row_cells, 2)   # col C (0-indexed = 2)
            location = _cell_val(row_cells, 3)      # col D
            manufacturer = _cell_val(row_cells, 4)  # col E
            ebif_uid = str(uid_val).strip()

            if manufacturer:
                complete += 1
            else:
                incomplete += 1

            # Keep first 100 rows for the expand view
            if len(rows_data) < 100:
                rows_data.append({
                    'ebif_uid': ebif_uid,
                    'tear_sheet': tear_sheet,
                    'location': location,
                    'manufacturer': manufacturer,
                })

        count = complete + incomplete
        details[sid] = {
            'count': count,
            'complete': complete,
            'incomplete': incomplete,
            'rows': rows_data,
        }

    wb.close()
    return details


def _cell_val(row_cells, col_idx):
    """Get a cell value by 0-indexed column, returning cleaned string or empty."""
    if col_idx >= len(row_cells):
        return ''
    val = row_cells[col_idx].value
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s in ('#REF!', '0') else s


def _find_sheet(wb, schedule_name):
    """Find sheet by name, handling emoji prefixes."""
    for sheet_name in wb.sheetnames:
        if sheet_name == schedule_name:
            return wb[sheet_name]
        stripped = sheet_name
        while stripped and (not stripped[0].isascii() or stripped[0] == ' '):
            stripped = stripped[1:]
        if stripped == schedule_name:
            return wb[sheet_name]
        if sheet_name.endswith(schedule_name):
            return wb[sheet_name]
    return None
