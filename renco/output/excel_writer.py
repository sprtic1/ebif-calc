"""EID-branded Excel output — 5 sheets per spec v03."""

import logging
import math
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..aggregator import ProjectTotals
from ..calculator import WallResult
from ..packing import PackingResult
from ..block_catalog import BlockCatalog
from ..wall_parser import Wall, _ft_in

logger = logging.getLogger(__name__)

# Styles
HDR_FILL = PatternFill(start_color="868D54", end_color="868D54", fill_type="solid")
BAND_FILL = PatternFill(start_color="C2C8A2", end_color="C2C8A2", fill_type="solid")
COM_FILL = PatternFill(start_color="868D54", end_color="868D54", fill_type="solid")
RES_FILL = PatternFill(start_color="C2C8A2", end_color="C2C8A2", fill_type="solid")
FLAG_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
EMPTY_FILL = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
HDR_FONT = Font(name="Arial Narrow", bold=True, color="FFFFFF", size=10)
BODY = Font(name="Arial Narrow", size=10)
BOLD = Font(name="Arial Narrow", bold=True, size=10)
TITLE = Font(name="Arial Narrow", bold=True, size=14)
WARN = Font(name="Arial Narrow", size=10, color="CC0000")
ITALIC = Font(name="Arial Narrow", size=9, italic=True, color="666666")
TOT_FONT = Font(name="Arial Narrow", bold=True, color="FFFFFF", size=10)
THIN = Border(bottom=Side(style="thin", color="AAAAAA"))
BOX = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))


def _aw(ws):
    for col_cells in ws.columns:
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(mx + 3, 50)


def _hdr(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _row_style(ws, row, ncols, banded):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY
        cell.border = THIN
        if banded:
            cell.fill = BAND_FILL


def write_excel(totals: ProjectTotals, results: list[WallResult], packing: PackingResult,
                catalog: BlockCatalog, output_path: str):
    wb = Workbook()

    # Determine block columns
    has_res = any(bid.startswith("RES") for bid in totals.blocks_by_type)
    def _block_sort_key(bid):
        """Sort by numeric suffix descending (largest block first)."""
        parts = bid.split("-")
        return -int(parts[1]) if len(parts) == 2 and parts[1].isdigit() else 0

    com_ids = sorted((bid for bid in catalog.all_ids() if bid.startswith("COM")), key=_block_sort_key)
    res_ids = sorted((bid for bid in catalog.all_ids() if bid.startswith("RES")), key=_block_sort_key) if has_res else []
    block_ids = com_ids + res_ids

    _write_summary(wb, totals, packing, catalog, block_ids)
    _write_wall_schedule(wb, results, catalog, block_ids)
    _write_container_layout(wb, packing)
    _write_container_manifest(wb, packing, catalog)
    _write_warnings(wb, totals, packing)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Excel: %s", output_path)


# === Sheet 1: Summary ===

def _write_summary(wb, totals, packing, catalog, block_ids):
    ws = wb.create_sheet("Summary", 0)
    ws.cell(1, 1, "RENCO Block Schedule").font = TITLE
    ws.cell(2, 1, f"Project: {totals.project_name}").font = BODY
    ws.cell(3, 1, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = BODY
    ws.cell(4, 1, f"Total Walls: {totals.total_walls}  |  Flagged: {totals.walls_with_flags}  |  Excluded: {totals.excluded_count}").font = BODY

    row = 6
    ncols = 6
    _hdr(ws, row, ["Block ID", "Series", "Count", "Weight/Block (lbs)", "Status", "Total Weight (lbs)"])
    row += 1

    ordered_bids = [bid for bid in block_ids if bid in totals.blocks_by_type]
    data_start = row
    for i, bid in enumerate(ordered_bids):
        cnt = totals.blocks_by_type[bid]
        b = catalog.get(bid)
        ws.cell(row, 1, bid)
        ws.cell(row, 2, "Residential" if bid.startswith("RES") else "Commercial")
        ws.cell(row, 3, cnt)
        ws.cell(row, 4, b["weight_lbs"])
        ws.cell(row, 5, "Confirmed" if b.get("confirmed_weight") else "Estimated").font = (
            BOLD if b.get("confirmed_weight") else ITALIC)
        ws.cell(row, 6, round(cnt * b["weight_lbs"], 1))
        _row_style(ws, row, ncols, i % 2 == 1)
        row += 1
    data_end = row - 1

    # Totals row
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = TOT_FONT; cell.fill = HDR_FILL; cell.border = THIN
    ws.cell(row, 1, "TOTAL")
    cl_count = get_column_letter(3)
    cl_weight = get_column_letter(6)
    ws.cell(row, 3, f"=SUM({cl_count}{data_start}:{cl_count}{data_end})")
    ws.cell(row, 6, f"=SUM({cl_weight}{data_start}:{cl_weight}{data_end})")
    row += 2

    ws.cell(row, 1, "(COM-16 weight confirmed. All others estimated proportional to volume.)").font = ITALIC
    row += 1
    ws.cell(row, 1, f"Containers: {packing.containers_required} x {packing.container_name}").font = BODY
    _aw(ws)


# === Sheet 2: Wall Schedule ===

def _write_wall_schedule(wb, results, catalog, block_ids):
    ws = wb.create_sheet("Wall Schedule")

    def _sk(wr):
        m = re.match(r"WALL-(\d+)(X?)", wr.wall.wall_id, re.IGNORECASE)
        return (int(m.group(1)), m.group(2)) if m else (9999, "")
    sorted_r = sorted(results, key=_sk)

    headers = ["Wall ID", "Story", "Height", "Wall Length", "Series", "Courses", "Openings"]
    headers += block_ids
    headers += ["Total Blocks", "Weight (lbs)"]
    ncols = len(headers)
    series_col = 5
    block_start = 8

    _hdr(ws, 1, headers)

    for i, wr in enumerate(sorted_r):
        w = wr.wall
        row = i + 2
        col = 1
        id_cell = ws.cell(row, col, w.wall_id); col += 1
        if w.flags:
            id_cell.fill = FLAG_FILL
        ws.cell(row, col, w.story_name); col += 1
        ws.cell(row, col, _ft_in(w.height_in)); col += 1
        ws.cell(row, col, _ft_in(w.length_in)); col += 1
        c = ws.cell(row, col, w.series.title()); c.alignment = Alignment(horizontal="right"); col += 1
        ws.cell(row, col, wr.courses); col += 1
        ws.cell(row, col, len(w.openings) or ""); col += 1

        for bid in block_ids:
            ws.cell(row, col, wr.blocks.get(bid, 0)); col += 1

        ws.cell(row, col, wr.total_blocks); col += 1
        ws.cell(row, col, wr.weight_lbs)

        _row_style(ws, row, ncols, i % 2 == 1)
        ws.cell(row, series_col).alignment = Alignment(horizontal="right")
        if w.flags:
            ws.cell(row, 1).fill = FLAG_FILL

    # Totals row
    tr = len(sorted_r) + 2
    for c in range(1, ncols + 1):
        cell = ws.cell(tr, c)
        cell.font = TOT_FONT; cell.fill = HDR_FILL; cell.border = THIN
    ws.cell(tr, 1, "TOTAL")
    # SUM formulas for numeric columns
    for c in [6, 7] + list(range(block_start, ncols + 1)):
        cl = get_column_letter(c)
        ws.cell(tr, c, f"=SUM({cl}2:{cl}{tr - 1})")

    ws.freeze_panes = "B2"
    _aw(ws)


# === Sheet 3: Container Layout ===

def _write_container_layout(wb, packing):
    ws = wb.create_sheet("Container Layout")
    ws.cell(1, 1, "Container Pallet Layout").font = TITLE
    ws.cell(2, 1, "Indicative layout — pallet heights and stacking TBC with Renco").font = ITALIC

    THICK = Side(style="medium", color="000000")
    THIN_SIDE = Side(style="thin", color="000000")

    def _cell_border(r_idx, c_idx, total_rows, total_cols):
        top = THICK if r_idx == 0 else THIN_SIDE
        bottom = THICK if r_idx == total_rows - 1 else THIN_SIDE
        left = THICK if c_idx == 0 else THIN_SIDE
        right = THICK if c_idx == total_cols - 1 else THIN_SIDE
        return Border(top=top, bottom=bottom, left=left, right=right)

    # Build per-container type mapping
    ctr_types = {}
    ctr_idx = 0
    if packing.recommended:
        for part in packing.recommended.containers:
            for _ in range(part["count"]):
                ctr_idx += 1
                ctr_types[ctr_idx] = part["type"]

    cr = 4
    pallets = packing.pallets

    for cnum in range(1, packing.containers_required + 1):
        ctr_type = ctr_types.get(cnum, {})
        ctr_ppl = ctr_type.get("pallets_per_layer", packing.pallets_per_layer)
        ctr_name = ctr_type.get("name", packing.container_name)
        # Grid: 40ft = 11 cols × 2 rows, 20ft = 5 cols × 2 rows
        num_cols = ctr_ppl // 2 if ctr_ppl > 1 else 1
        num_rows = 2

        for c in range(1, num_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14

        ctr_pallets = [p for p in pallets if p.container == cnum]
        layers_used = max((p.layer for p in ctr_pallets), default=1)

        # Header
        ctr_wt = sum(p.weight_lbs for p in ctr_pallets)
        max_wt = ctr_type.get("max_payload_lbs", 0)
        max_pal = ctr_type.get("max_pallets", 0)
        ws.cell(cr, 1, f"Container {cnum} of {packing.containers_required} — {ctr_name}").font = Font(
            name="Arial Narrow", bold=True, size=12)
        cr += 1
        ws.cell(cr, 1,
            f"Pallets: {len(ctr_pallets)}/{max_pal}  |  "
            f"Weight: {ctr_wt:,.0f}/{max_wt:,.0f} lbs  |  "
            f"{num_cols} cols × {num_rows} rows per layer"
        ).font = ITALIC
        cr += 1

        total_grid_rows = layers_used * num_rows

        # Render grid per layer: row 1 then row 2, col 1..N
        # Packing assigns p.row = 1 or 2 (width), p.col = 1..N (columns)
        for layer in range(1, layers_used + 1):
            layer_pallets = {(p.row, p.col): p for p in ctr_pallets if p.layer == layer}
            for r in range(1, num_rows + 1):
                grid_r_idx = (layer - 1) * num_rows + (r - 1)
                ws.row_dimensions[cr].height = 45
                for c in range(1, num_cols + 1):
                    cell = ws.cell(cr, c)
                    p = layer_pallets.get((r, c))
                    if p:
                        cell.value = f"P{p.pallet_number}{chr(10)}{p.block_id}{chr(10)}{p.block_count}x"
                        cell.fill = COM_FILL if p.block_id.startswith("COM") else RES_FILL
                        cell.font = Font(name="Arial Narrow", size=8, color="FFFFFF", bold=True)
                    else:
                        cell.value = "EMPTY"
                        cell.fill = EMPTY_FILL
                        cell.font = Font(name="Arial Narrow", size=8, color="666666")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = _cell_border(grid_r_idx, c - 1, total_grid_rows, num_cols)
                cr += 1

            if layer < layers_used:
                ws.cell(cr - num_rows, num_cols + 1, f"← Layer {layer}").font = ITALIC
        ws.cell(cr - num_rows, num_cols + 1, f"← Layer {layers_used}").font = ITALIC

        cr += 3

    # Legend
    ws.cell(cr, 1, "Legend:").font = BOLD; cr += 1
    c = ws.cell(cr, 1, "  Commercial (COM)"); c.fill = COM_FILL; c.font = Font(name="Arial Narrow", size=9, color="FFFFFF", bold=True); cr += 1
    c = ws.cell(cr, 1, "  Residential (RES)"); c.fill = RES_FILL; c.font = Font(name="Arial Narrow", size=9, bold=True); cr += 1
    c = ws.cell(cr, 1, "  Empty"); c.fill = EMPTY_FILL; c.font = Font(name="Arial Narrow", size=9, color="666666")


# === Sheet 4: Container Manifest ===

def _write_container_manifest(wb, packing, catalog):
    ws = wb.create_sheet("Container Manifest")
    ws.cell(1, 1, "Container Manifest").font = TITLE

    # Recommended combination
    rec = packing.recommended
    ws.cell(2, 1, f"Recommended: {rec.label}").font = Font(name="Arial Narrow", bold=True, size=11, color="868D54")
    ws.cell(3, 1, f"Total Pallets: {packing.total_pallets}  |  Utilization: {rec.utilization_percent}%").font = BODY

    # Pallet breakdown by block type
    row = 5
    _hdr(ws, row, ["Block ID", "Total Blocks", "Blocks/Pallet", "Pallets", "Weight/Block", "Pallet Weight", "Dims"])
    row += 1

    by_type: dict[str, dict] = {}
    for p in packing.pallets:
        if p.block_id not in by_type:
            by_type[p.block_id] = {"blocks": 0, "pallets": 0}
        by_type[p.block_id]["pallets"] += 1
        by_type[p.block_id]["blocks"] += p.block_count

    total_wt = 0.0
    for i, bid in enumerate(sorted(by_type)):
        info = by_type[bid]
        b = catalog.get(bid)
        wpb = b["weight_lbs"]
        spec = packing.pallet_specs.get(bid)
        bpp = spec.blocks_per_pallet if spec else packing.blocks_per_pallet
        pwt = round(bpp * wpb, 1)
        total_wt += info["pallets"] * pwt
        ws.cell(row, 1, bid)
        ws.cell(row, 2, info["blocks"])
        ws.cell(row, 3, bpp)
        ws.cell(row, 4, info["pallets"])
        ws.cell(row, 5, wpb)
        ws.cell(row, 6, pwt)
        h = f"{spec.pallet_height_in:.0f}" if spec else "48"
        ws.cell(row, 7, f"48x40x{h}in")
        _row_style(ws, row, 7, i % 2 == 1)
        row += 1

    # Total weight row
    for c in range(1, 8):
        cell = ws.cell(row, c); cell.font = TOT_FONT; cell.fill = HDR_FILL; cell.border = THIN
    ws.cell(row, 1, "TOTAL")
    ws.cell(row, 4, packing.total_pallets)
    ws.cell(row, 6, round(total_wt, 1))
    row += 2

    # Consumables section
    cons = packing.consumables
    ws.cell(row, 1, "Consumables & Equipment").font = Font(name="Arial Narrow", bold=True, size=11)
    row += 1
    _hdr(ws, row, ["Item", "Quantity", "Unit", "Notes"])
    row += 1
    est = " (estimated)" if not cons.adhesive_confirmed else ""
    ws.cell(row, 1, "Adhesive (Plexus MA530)")
    ws.cell(row, 2, cons.adhesive_cartridges)
    ws.cell(row, 3, "cartridges")
    ws.cell(row, 4, f"{cons.total_joint_linear_ft:.0f} LF joints / {cons.adhesive_per_cartridge_ft:.0f} LF per cart{est}")
    _row_style(ws, row, 4, False)
    row += 1
    ws.cell(row, 1, "Rubber Mallets")
    ws.cell(row, 2, cons.mallets)
    ws.cell(row, 3, "units")
    ws.cell(row, 4, f"Crew size: {cons.crew_size}")
    _row_style(ws, row, 4, True)
    row += 2

    # Single-container analysis
    REC_FILL = PatternFill(start_color="E8EBDA", end_color="E8EBDA", fill_type="solid")
    FAIL_FONT = Font(name="Arial Narrow", size=9, color="999999")

    singles = getattr(packing, '_single_options', [])
    if singles:
        ws.cell(row, 1, "Single-Container Analysis").font = Font(name="Arial Narrow", bold=True, size=11)
        row += 1
        _hdr(ws, row, ["Container", "Pallets", "Payload", "Height", "Status"])
        row += 1
        for s in singles:
            ctr = s["type"]
            ws.cell(row, 1, ctr["name"])
            ws.cell(row, 2, f"{ctr['max_pallets']} max")
            ws.cell(row, 3, f"{ctr['max_payload_lbs']:,.0f} lbs")
            ws.cell(row, 4, f"{ctr['interior_height_in']}in")
            ws.cell(row, 5, s["reason"])
            if s["works"]:
                for c in range(1, 6):
                    ws.cell(row, c).fill = REC_FILL
                    ws.cell(row, c).font = Font(name="Arial Narrow", bold=True, size=10)
            else:
                ws.cell(row, 5).font = FAIL_FONT
            _row_style(ws, row, 5, False)
            row += 1
        row += 1

    # All container options
    ws.cell(row, 1, "Container Options Evaluated").font = Font(name="Arial Narrow", bold=True, size=11)
    row += 1
    _hdr(ws, row, ["", "Combination", "Containers", "Capacity", "Utilization", "Payload"])
    row += 1

    for i, opt in enumerate(packing.all_options):
        tag = "RECOMMENDED" if opt.is_recommended else ""
        ws.cell(row, 1, tag)
        if opt.is_recommended:
            ws.cell(row, 1).font = Font(name="Arial Narrow", bold=True, size=10, color="868D54")
        ws.cell(row, 2, opt.label)
        ws.cell(row, 3, opt.total_containers)
        ws.cell(row, 4, opt.total_capacity)
        ws.cell(row, 5, f"{opt.utilization_percent}%")
        ws.cell(row, 6, f"{opt.total_payload_lbs:,.0f}")
        _row_style(ws, row, 6, i % 2 == 1)
        if opt.is_recommended:
            for c in range(1, 7):
                ws.cell(row, c).fill = REC_FILL
                ws.cell(row, c).font = Font(name="Arial Narrow", bold=True, size=10)
        row += 1

    _aw(ws)


# === Sheet 5: Assumptions & Warnings ===

def _write_warnings(wb, totals, packing):
    ws = wb.create_sheet("Assumptions & Warnings")
    ws.cell(1, 1, "Assumptions & Warnings").font = TITLE

    row = 3
    _hdr(ws, row, ["#", "Category", "Detail"]); row += 1

    assumptions = [
        ("Catalog", "Block dimensions from RENCO Sales Brochure 2024, page 12"),
        ("Catalog", "COM-16 weight (8.5 lbs) confirmed. All others estimated proportional to volume."),
        ("Packing", f"Pallet specs assumed: {packing.blocks_per_pallet} blocks/pallet — TBC from Renco"),
        ("Calculation", "Running bond pattern assumed for all walls"),
        ("Calculation", "No waste factor applied"),
        ("Filter", "Only walls with composite name containing 'renco' are processed"),
        ("Geometry", "Curved walls (arcAngle > 0) and known irregular walls excluded"),
    ]
    idx = 1
    for cat, det in assumptions:
        ws.cell(row, 1, idx); ws.cell(row, 2, cat); ws.cell(row, 3, det)
        _row_style(ws, row, 3, idx % 2 == 0)
        row += 1; idx += 1

    if totals.warnings:
        row += 1
        ws.cell(row, 1, "Wall-Level Warnings").font = Font(name="Arial Narrow", bold=True, size=11)
        row += 1
        for w in totals.warnings:
            ws.cell(row, 1, idx); ws.cell(row, 2, "Wall"); ws.cell(row, 3, w)
            ws.cell(row, 3).font = WARN
            _row_style(ws, row, 3, idx % 2 == 0)
            row += 1; idx += 1

    if packing.warnings:
        row += 1
        ws.cell(row, 1, "Packing Warnings").font = Font(name="Arial Narrow", bold=True, size=11)
        row += 1
        for w in packing.warnings:
            ws.cell(row, 1, idx); ws.cell(row, 2, "Packing"); ws.cell(row, 3, w)
            _row_style(ws, row, 3, idx % 2 == 0)
            row += 1; idx += 1

    _aw(ws)
