"""Excel Reader — reads a completed EID schedule workbook back into pipeline data.

Used in Step 2 (Publish) to read the working document after manual data entry
and convert it back into structured schedule data for the website dashboard.
"""

import logging
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def read_workbook(
    filepath: Path,
    schedule_defs: list[dict],
) -> dict[str, list[dict]]:
    """Read an EID schedule Excel workbook and return structured schedule data.

    Args:
        filepath: Path to the .xlsx file
        schedule_defs: Schedule definitions from schedules.json

    Returns:
        Dict mapping schedule_id -> list of row dicts (same format as extractor output).
    """
    wb = load_workbook(str(filepath), data_only=True)
    logger.info("Reading workbook: %s (%d sheets)", filepath.name, len(wb.sheetnames))

    # Build a lookup from tab name -> schedule def
    tab_to_def: dict[str, dict] = {}
    for sdef in schedule_defs:
        tab_name = sdef["name"][:31]  # Excel max 31 chars
        tab_to_def[tab_name] = sdef

    schedules: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        # Skip non-schedule tabs
        if sheet_name in ("Summary", "QC Audit"):
            continue

        sdef = tab_to_def.get(sheet_name)
        if not sdef:
            logger.debug("Skipping unknown sheet: %s", sheet_name)
            continue

        ws = wb[sheet_name]
        rows = _read_sheet(ws)
        schedules[sdef["id"]] = rows
        logger.info("Read '%s': %d rows", sheet_name, len(rows))

    # Initialize empty lists for schedules not found in workbook
    for sdef in schedule_defs:
        if sdef["id"] not in schedules:
            schedules[sdef["id"]] = []

    total = sum(len(v) for v in schedules.values())
    active = sum(1 for v in schedules.values() if v)
    logger.info("Read %d elements across %d active schedules", total, active)
    return schedules


def _read_sheet(ws) -> list[dict]:
    """Read a single worksheet into a list of row dicts.

    Assumes row 1 is the header row, data starts at row 2.
    """
    rows = list(ws.iter_rows(min_row=1, values_only=False))
    if not rows:
        return []

    # Extract header names from row 1
    header_row = rows[0]
    headers = []
    for cell in header_row:
        val = cell.value
        if val is not None:
            headers.append(str(val).strip())
        else:
            headers.append("")

    if not any(headers):
        return []

    # Read data rows
    result = []
    for row in rows[1:]:
        entry = {}
        all_empty = True
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                val = cell.value
                if val is not None:
                    val = str(val).strip() if not isinstance(val, (int, float)) else val
                    all_empty = False
                else:
                    val = ""
                entry[headers[col_idx]] = val

        # Skip completely empty rows
        if not all_empty:
            # Add internal fields for compatibility
            entry.setdefault("_guid", "")
            entry.setdefault("_type", "")
            entry.setdefault("Qty", 1)
            result.append(entry)

    return result


def find_workbook(output_dir: Path, project_slug: str) -> Path | None:
    """Find the schedule workbook for a project.

    Looks for ebif_schedule_{slug}.xlsx in the output directory.
    Returns the path or None if not found.
    """
    expected = output_dir / f"ebif_schedule_{project_slug}.xlsx"
    if expected.exists():
        return expected

    # Fallback: look for any ebif_schedule_*.xlsx
    candidates = sorted(output_dir.glob("ebif_schedule_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if candidates:
        logger.info("Using most recent workbook: %s", candidates[0].name)
        return candidates[0]

    return None
