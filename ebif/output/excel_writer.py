"""EID-Branded Excel Writer — generates individual Excel files per schedule.

Each active schedule gets its own file: Appliances.xlsx, Furniture.xlsx, etc.
Plus a Summary.xlsx with overview and a QC Audit.xlsx for publish mode.

Uses EID brand colors: Olive #868C54, Sage #C2C8A2, Warm Gray #737569.
"""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# EID Brand Colors
OLIVE = "868C54"
SAGE = "C2C8A2"
WARM_GRAY = "737569"
WHITE = "FFFFFF"
LIGHT_YELLOW = "FFFFF0"

# Fonts
HEADER_FONT = Font(name="Lato", bold=True, color=WHITE, size=10)
BODY_FONT = Font(name="Arial Narrow", color="2C2C2C", size=10)
TITLE_FONT = Font(name="Lato", bold=True, color=OLIVE, size=14)
SUBTITLE_FONT = Font(name="Lato", color=WARM_GRAY, size=11)

# Fills
HEADER_FILL = PatternFill(start_color=OLIVE, end_color=OLIVE, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=SAGE, end_color=SAGE, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _set_column_widths(ws):
    """Set all column widths to 200 pixels (~28.5 characters in openpyxl units)."""
    # openpyxl width is in characters; 200px ~ 28.5 chars at default font
    COL_WIDTH = 28.5
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTH


def _write_header_row(ws, row: int, columns: list[str]):
    """Write a styled header row."""
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_data_rows(ws, start_row: int, rows: list[dict], columns: list[str]):
    """Write data rows with alternating color."""
    for i, data_row in enumerate(rows):
        row_num = start_row + i
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        for col_idx, col_name in enumerate(columns, start=1):
            val = data_row.get(col_name, "")
            if val is None:
                val = ""
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _count_fill_status(rows: list[dict], columns: list[str]) -> tuple[int, int]:
    """Count populated vs blank fields across all rows and columns."""
    populated = 0
    blank = 0
    skip = {"Element ID", "Qty", "Zone", "Room"}
    for row in rows:
        for col in columns:
            if col in skip:
                continue
            val = row.get(col, "")
            if val is not None and str(val).strip() not in ("", "None", "0"):
                populated += 1
            else:
                blank += 1
    return populated, blank


def _schedule_filename(schedule_name: str) -> str:
    """Convert schedule name to a clean filename."""
    return f"{schedule_name}.xlsx"


# ------------------------------------------------------------------
# Individual schedule file
# ------------------------------------------------------------------

def write_schedule_file(
    schedule_def: dict,
    rows: list[dict],
    output_dir: Path,
    project_name: str,
) -> Path:
    """Write a single schedule to its own Excel file.

    Returns the path to the written file.
    """
    sname = schedule_def["name"]
    wb = Workbook()
    ws = wb.active
    ws.title = sname[:31]

    # Title row
    ws.cell(row=1, column=1, value=sname).font = TITLE_FONT
    ws.cell(row=2, column=1, value=project_name).font = Font(name="Arial Narrow", color=WARM_GRAY, size=10)
    ws.cell(row=3, column=1, value=f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(
        name="Arial Narrow", color=WARM_GRAY, size=9, italic=True)

    # Data table starts at row 5
    # Archicad GUID is the last column — hidden but present as the primary key
    columns = ["Element ID"] + schedule_def.get("columns", []) + ["Qty", "Archicad GUID"]
    _write_header_row(ws, 5, columns)

    # Write data — map _guid to "Archicad GUID" column
    for i, data_row in enumerate(rows):
        row_num = 6 + i
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        for col_idx, col_name in enumerate(columns, start=1):
            if col_name == "Archicad GUID":
                val = data_row.get("_guid", "")
            else:
                val = data_row.get(col_name, "")
            if val is None:
                val = ""
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _set_column_widths(ws)

    # Hide the Archicad GUID column (last column)
    guid_col = get_column_letter(len(columns))
    ws.column_dimensions[guid_col].hidden = True

    # Freeze header row
    ws.freeze_panes = "A6"

    # Auto-filter
    if rows:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A5:{last_col}{len(rows) + 5}"

    # Save
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = _schedule_filename(sname)
    filepath = output_dir / filename
    wb.save(str(filepath))
    logger.info("Schedule file: %s (%d rows)", filename, len(rows))
    return filepath


# ------------------------------------------------------------------
# Summary file
# ------------------------------------------------------------------

def write_summary_file(
    schedules: dict[str, list[dict]],
    schedule_defs: list[dict],
    project_name: str,
    output_dir: Path,
    qc_issues: list[dict] | None = None,
    mode: str = "template",
) -> Path:
    """Write the Summary.xlsx overview file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.cell(row=1, column=1, value="EBIF-CALC").font = Font(name="Lato", bold=True, color=OLIVE, size=20)
    ws.cell(row=2, column=1, value="Ellis Building Intelligence Framework").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value=project_name).font = Font(name="Arial Narrow", color=WARM_GRAY, size=11)

    mode_label = "Working Document (Step 1 -- fill in spec data)" if mode == "template" else "Published Report (Step 2)"
    ws.cell(row=4, column=1, value=mode_label).font = Font(
        name="Arial Narrow", color=OLIVE, size=10, bold=True, italic=True)
    ws.cell(row=5, column=1, value=f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(
        name="Arial Narrow", color=WARM_GRAY, size=9, italic=True)

    # Table
    row = 7
    if mode == "template":
        headers = ["#", "Schedule", "File", "Elements", "Fields Populated", "Fields Blank"]
    else:
        headers = ["#", "Schedule", "File", "Elements", "Warnings", "Complete"]
    _write_header_row(ws, row, headers)

    row = 8
    total_elements = 0
    for idx, sdef in enumerate(schedule_defs, start=1):
        sid = sdef["id"]
        sched_rows = schedules.get(sid, [])
        count = len(sched_rows)
        total_elements += count
        filename = _schedule_filename(sdef["name"]) if count > 0 else "--"

        if mode == "template":
            pop, blank = _count_fill_status(sched_rows, sdef.get("columns", []))
            vals = [idx, sdef["name"], filename, count, pop, blank]
        else:
            warnings = sum(1 for q in (qc_issues or []) if q["schedule"] == sdef["name"] and q["severity"] == "Warning")
            complete = count - warnings if count > 0 else 0
            vals = [idx, sdef["name"], filename, count, warnings, complete]

        fill = ALT_ROW_FILL if idx % 2 == 0 else WHITE_FILL
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
        row += 1

    for col_idx, val in enumerate(["", "TOTAL", "", total_elements, "", ""], start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name="Lato", bold=True, color=OLIVE, size=10)
        cell.border = THIN_BORDER

    _set_column_widths(ws)

    output_dir.mkdir(parents=True, exist_ok=True)
    filepath = output_dir / "Summary.xlsx"
    wb.save(str(filepath))
    logger.info("Summary file: Summary.xlsx (%d schedules, %d elements)", len(schedule_defs), total_elements)
    return filepath


# ------------------------------------------------------------------
# QC Audit file
# ------------------------------------------------------------------

def write_qc_file(
    qc_issues: list[dict],
    output_dir: Path,
) -> Path:
    """Write the QC Audit.xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "QC Audit"

    columns = ["Schedule", "Element ID", "Severity", "Message"]
    _write_header_row(ws, 1, columns)

    for i, issue in enumerate(qc_issues):
        row_num = 2 + i
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        for col_idx, key in enumerate(["schedule", "element_id", "severity", "message"], start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=issue.get(key, ""))
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER

    _set_column_widths(ws)
    ws.freeze_panes = "A2"

    output_dir.mkdir(parents=True, exist_ok=True)
    filepath = output_dir / "QC Audit.xlsx"
    wb.save(str(filepath))
    logger.info("QC file: QC Audit.xlsx (%d issues)", len(qc_issues))
    return filepath


# ------------------------------------------------------------------
# Step 1: Write all template files
# ------------------------------------------------------------------

def write_template(
    schedules: dict[str, list[dict]],
    schedule_defs: list[dict],
    project_name: str,
    output_dir: Path,
) -> list[Path]:
    """Generate Step 1 individual Excel files (working documents).

    Returns list of paths written.
    """
    paths = []

    # Summary
    summary_path = write_summary_file(schedules, schedule_defs, project_name, output_dir, mode="template")
    paths.append(summary_path)

    # Individual schedule files
    for sdef in schedule_defs:
        sid = sdef["id"]
        rows = schedules.get(sid, [])
        if rows:
            p = write_schedule_file(sdef, rows, output_dir, project_name)
            paths.append(p)

    logger.info("Step 1: wrote %d files to %s", len(paths), output_dir)
    return paths


# ------------------------------------------------------------------
# Step 2: Write all published files
# ------------------------------------------------------------------

def write_published(
    schedules: dict[str, list[dict]],
    schedule_defs: list[dict],
    qc_issues: list[dict],
    project_name: str,
    output_dir: Path,
) -> list[Path]:
    """Generate Step 2 published Excel files with QC.

    Returns list of paths written.
    """
    paths = []

    # Summary (publish mode)
    summary_path = write_summary_file(schedules, schedule_defs, project_name, output_dir,
                                       qc_issues=qc_issues, mode="publish")
    paths.append(summary_path)

    # Individual schedule files (overwrite with latest data)
    for sdef in schedule_defs:
        sid = sdef["id"]
        rows = schedules.get(sid, [])
        if rows:
            p = write_schedule_file(sdef, rows, output_dir, project_name)
            paths.append(p)

    # QC file
    if qc_issues:
        qc_path = write_qc_file(qc_issues, output_dir)
        paths.append(qc_path)

    logger.info("Step 2: wrote %d files to %s", len(paths), output_dir)
    return paths
