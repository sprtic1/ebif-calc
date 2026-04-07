"""Merge — preserves manually entered spec data across EBIF Report re-runs.

When 'Run the EBIF Report' runs on an existing project, this module reads
the existing Excel files, matches rows by Archicad GUID, and merges:
- Archicad data (Element ID, toggle status) is always refreshed
- Manually entered spec data (TEAR SHEET #, MANUFACTURER, etc.) is preserved
  if the Excel has values that Archicad doesn't
"""

import logging
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Columns that always come from Archicad (never preserve from Excel)
ARCHICAD_OWNED = {"EBIF UID", "Element ID", "_guid", "_type", "Qty"}


def load_existing_data(filepath: Path) -> dict[str, dict]:
    """Load existing Excel data keyed by Archicad GUID.

    Returns dict mapping GUID -> row dict.
    """
    if not filepath.exists():
        return {}

    try:
        wb = load_workbook(str(filepath), data_only=True)
        ws = wb.active
    except Exception as e:
        logger.warning("Could not read %s for merge: %s", filepath.name, e)
        return {}

    # Find header row
    header_row = None
    for r in range(1, 10):
        for c in range(1, 30):
            if ws.cell(r, c).value and str(ws.cell(r, c).value).strip() == "Element ID":
                header_row = r
                break
        if header_row:
            break

    if not header_row:
        return {}

    # Extract headers
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            headers.append(str(v).strip())
        else:
            headers.append("")

    # Find primary key column: prefer EBIF UID, fall back to Archicad GUID
    key_col = None
    key_name = None
    for preferred in ("EBIF UID",):
        for i, h in enumerate(headers):
            if h == preferred:
                key_col = i
                key_name = preferred
                break
        if key_col is not None:
            break

    if key_col is None:
        logger.info("No EBIF UID column in %s -- cannot merge", filepath.name)
        return {}

    # Read rows keyed by the primary key
    existing = {}
    for r in range(header_row + 1, ws.max_row + 1):
        guid = ws.cell(r, key_col + 1).value
        if not guid or not str(guid).strip():
            continue
        guid = str(guid).strip()
        row = {}
        for ci, h in enumerate(headers):
            if h:
                val = ws.cell(r, ci + 1).value
                row[h] = val if val is not None else ""
        existing[guid] = row

    logger.info("Loaded %d existing rows from %s for merge (key: %s)", len(existing), filepath.name, key_name)
    return existing


def merge_schedules(
    fresh_data: dict[str, list[dict]],
    schedule_defs: list[dict],
    output_dir: Path,
) -> dict[str, list[dict]]:
    """Merge fresh Archicad data with existing Excel data.

    For each schedule:
    1. Load existing Excel file (if it exists)
    2. Match rows by Archicad GUID
    3. For matched rows: keep Archicad-owned fields fresh, preserve manual entries
    4. New rows (not in existing): add as-is
    5. Deleted rows (in existing but not in fresh): drop them

    Returns the merged schedule data.
    """
    merged = {}

    for sdef in schedule_defs:
        sid = sdef["id"]
        fresh_rows = fresh_data.get(sid, [])
        if not fresh_rows:
            merged[sid] = fresh_rows
            continue

        # Try to load existing data
        filepath = output_dir / f"{sdef['name']}.xlsx"
        existing = load_existing_data(filepath)

        if not existing:
            merged[sid] = fresh_rows
            continue

        # Merge
        result = []
        preserved_count = 0
        for row in fresh_rows:
            # Match by EBIF UID first, fall back to Archicad GUID
            key = row.get("EBIF UID", "") or row.get("_guid", "")
            if key and key in existing:
                old_row = existing[key]
                # Start with fresh Archicad data
                merged_row = dict(row)
                # For each column, if fresh value is empty but old has data, preserve it
                for key, old_val in old_row.items():
                    if key in ARCHICAD_OWNED:
                        continue
                    fresh_val = merged_row.get(key, "")
                    old_str = str(old_val).strip() if old_val is not None else ""
                    fresh_str = str(fresh_val).strip() if fresh_val is not None else ""
                    if not fresh_str and old_str and old_str != "0":
                        merged_row[key] = old_val
                        preserved_count += 1
                result.append(merged_row)
            else:
                result.append(row)

        merged[sid] = result
        if preserved_count > 0:
            logger.info("Merged '%s': preserved %d manual entries across %d rows",
                        sdef["name"], preserved_count, len(result))

    return merged
