"""Tests for excel_writer — generates workbook from mock data."""

import json
import tempfile
from pathlib import Path

import pytest

MOCK_DIR = Path(__file__).parent / "mock_data"


def load_mock():
    with open(MOCK_DIR / "sample_elements.json") as f:
        return json.load(f)


def load_schedule_defs():
    config_dir = Path(__file__).parent.parent / "config"
    with open(config_dir / "schedules.json") as f:
        return json.load(f)["schedules"]


def test_write_template():
    from ebif.output.excel_writer import write_template

    data = load_mock()
    defs = load_schedule_defs()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = write_template(
            schedules=data["schedules"],
            schedule_defs=defs,
            project_name="Test Project",
            output_dir=Path(tmpdir),
            project_slug="test-project",
        )
        assert path.exists()
        assert path.suffix == ".xlsx"
        assert path.stat().st_size > 0
        assert "test-project" in path.name


def test_write_published():
    from ebif.output.excel_writer import write_published

    data = load_mock()
    defs = load_schedule_defs()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = write_published(
            schedules=data["schedules"],
            schedule_defs=defs,
            qc_issues=[{"schedule": "Furniture", "element_id": "OBJ-011",
                         "guid": "fff-222", "severity": "Warning",
                         "message": "Missing TEAR SHEET #"}],
            project_name="Test Project",
            output_dir=Path(tmpdir),
            project_slug="test-project",
        )
        assert path.exists()
        assert "published" in path.name


def test_template_has_schedule_tabs():
    from openpyxl import load_workbook
    from ebif.output.excel_writer import write_template

    data = load_mock()
    defs = load_schedule_defs()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = write_template(
            schedules=data["schedules"],
            schedule_defs=defs,
            project_name="Test Project",
            output_dir=Path(tmpdir),
            project_slug="test-project",
        )
        wb = load_workbook(str(path))
        assert "Summary" in wb.sheetnames
        assert len(wb.sheetnames) >= 2


def test_roundtrip_excel():
    """Write template, read it back, verify data integrity."""
    from ebif.output.excel_writer import write_template
    from ebif.excel_reader import read_workbook

    data = load_mock()
    defs = load_schedule_defs()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = write_template(
            schedules=data["schedules"],
            schedule_defs=defs,
            project_name="Test Project",
            output_dir=Path(tmpdir),
            project_slug="test-project",
        )
        # Read it back
        result = read_workbook(path, defs)
        # Check furniture came back
        assert "furniture" in result
        assert len(result["furniture"]) == len(data["schedules"]["furniture"])
